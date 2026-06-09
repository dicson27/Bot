import sys
import io
import os # IMPORT ADICIONADO PARA LER VARIÁVEIS DA NUVEM

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import asyncio
import requests
import pandas as pd
from datetime import datetime, timedelta
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule

# Imports para o Servidor de Integração Direta com o Painel
from http.server import BaseHTTPRequestHandler, HTTPServer
import threading
from urllib.parse import urlparse, parse_qs

# =============================================
# CONFIGURAÇÕES
# =============================================
# TOKENS PROTEGIDOS (Puxa da nuvem ou usa os valores fixos localmente)
TOKEN_TELEGRAM = os.environ.get("TOKEN_TELEGRAM", "8657147116:AAFBvWsufUo_smhPveTsog0IiUoim4tJTYc")
CHAT_ID = os.environ.get("CHAT_ID", "938715543")

DURACAO = {
    "Futebol": 105, 
    "Basquete": 150, 
    "Tenis": 120, 
    "Hoquei": 150,
    "Futebol Americano": 190,
    "Beisebol": 180
}

LIGAS_PRIORITARIAS = [
    "premier league", "serie a", "brasileirao", "brasileirão",
    "copa do brasil", "libertadores", "sul-americana",
    "nba", "euroleague", "euroliga", "nbb",
    "atp", "wta", "itf", "challenger",
    "nhl", "shl", "del", "segunda divisao", "serie b",
    "nfl", "mlb"
]

URLS_ESPORTES = {
    "Futebol":  "https://www.flashscore.com.br/futebol/",
    "Basquete": "https://www.flashscore.com.br/basquete/",
    "Tenis":    "https://www.flashscore.com.br/tenis/",
    "Hoquei":   "https://www.flashscore.com.br/hoquei/", 
    "Futebol Americano": "https://www.flashscore.com.br/futebol-americano/",
    "Beisebol": "https://www.flashscore.com.br/beisebol/"
}

o_loop = None # Guardará o loop do asyncio

# =============================================
# 1. SCRAPING ROBUSTO
# =============================================
async def rolar_pagina_completa(page):
    print("   📜 Rolando página completa...")
    altura_anterior = 0
    tentativas_sem_mudanca = 0
    while tentativas_sem_mudanca < 3:
        altura_atual = await page.evaluate("document.body.scrollHeight")
        if altura_atual == altura_anterior: tentativas_sem_mudanca += 1
        else: tentativas_sem_mudanca = 0
        altura_anterior = altura_atual
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(1500)
    await page.evaluate("window.scrollTo(0, 0)")
    await page.wait_for_timeout(1000)

async def navegar_para_data(page, data_alvo: datetime):
    hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    alvo = data_alvo.replace(hour=0, minute=0, second=0, microsecond=0)
    diff_dias = (alvo - hoje).days
    if diff_dias == 0:
        print("   📅 Data já é hoje.")
        return
    print(f"   📅 Navegando para {data_alvo.strftime('%d/%m/%Y')} ({diff_dias:+d} dias)...")
    await page.keyboard.press("Escape")
    await page.wait_for_timeout(500)
    seta = "button[data-day-picker-arrow='next']" if diff_dias > 0 else "button[data-day-picker-arrow='prev']"
    for i in range(abs(diff_dias)):
        try:
            await page.click(seta, timeout=5000)
            await page.wait_for_timeout(1000)
        except Exception as e:
            print(f"   ⚠️ Falha na seta: {e}")
            break

async def extrair_texto_esporte(page, esporte, url_base, data: datetime):
    print(f"   🌐 Acessando {esporte} → {url_base}")
    try:
        await page.goto(url_base, wait_until="domcontentloaded", timeout=60000)
        await page.wait_for_timeout(5000)
        try:
            await page.click("button#onetrust-accept-btn-handler", timeout=3000)
            await page.wait_for_timeout(1000)
        except: pass
        
        await navegar_para_data(page, data)
        await page.wait_for_timeout(3000)
        
        hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        alvo = data.replace(hour=0, minute=0, second=0, microsecond=0)
        if alvo > hoje:
            for seletor_aba in ["button:has-text('PRÓXIMOS')", "text=PRÓXIMOS", "button:has-text('TODOS')", "text=TODOS"]:
                try:
                    await page.click(seletor_aba, timeout=3000)
                    await page.wait_for_timeout(2000)
                    break
                except: continue
        elif alvo < hoje:
            for seletor_aba in ["button:has-text('ENCERRADOS')", "text=ENCERRADOS"]:
                try:
                    await page.click(seletor_aba, timeout=3000)
                    await page.wait_for_timeout(2000)
                    break
                except: continue
        else:
            for seletor_aba in ["button:has-text('TODOS')", "text=TODOS"]:
                try:
                    await page.click(seletor_aba, timeout=3000)
                    await page.wait_for_timeout(2000)
                    break
                except: continue
                
        await rolar_pagina_completa(page)
        texto = await page.evaluate("document.body.innerText")
        print(f"   ✅ {esporte} coletado! ({len(texto)} caracteres)")
        return texto
    except Exception as e:
        print(f"   ⚠️ Erro em {esporte}: {e}")
        return ""

async def extrair_jogos_flashscore(data: datetime = None):
    if data is None: data = datetime.now()
    print(f"\n🚀 Iniciando coleta para {data.strftime('%d/%m/%Y')}...")
    dados_por_esporte = {}
    async with async_playwright() as p:
        # NAVEGADOR CONFIGURADO PARA RODAR INVISÍVEL NA NUVEM (headless=True) E SEM SANDBOX
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"])
        context = await browser.new_context(viewport={"width": 1366, "height": 768})
        page = await context.new_page()
        for esporte, url in URLS_ESPORTES.items():
            texto = await extrair_texto_esporte(page, esporte, url, data)
            if texto: dados_por_esporte[esporte] = texto
        await browser.close()
    return dados_por_esporte

# =============================================
# 2. EXTRAÇÃO DIRETA
# =============================================
def extrair_jogos_esporte_direto(esporte, texto):
    import re
    SKIP = {
        'PREVIEW', 'SRF', 'Classificação', 'Tabela', 'Classificação ao vivo',
        'TODOS', 'AO VIVO', 'ODDS', 'ENCERRADOS', 'PRÓXIMOS', 'Publicidade',
        'LIGAS FIXADAS', 'MINHAS EQUIPES', 'ADICIONAR EQUIPE', 'RANKINGS',
        'FAVORITOS', 'FUTEBOL', 'BASQUETE', 'TÊNIS', 'VÔLEI', 'FUTEBOL AM.',
        'BEISEBOL', 'HANDEBOL', 'ACESSAR', 'RESULTADOS', 'NOTÍCIAS', 'APOSTAS',
        'FIFA', 'PAÍSES', 'Mostrar mais', '-', 'CALENDÁRIO', 'CATEGORIAS', 'TORNEIOS ATUAIS',
    }
    time_re  = re.compile(r'^\d{2}:\d{2}$')
    pais_re  = re.compile(r'^[A-ZÁÉÍÓÚÀÃÕÂÊÎÔÛÇ][A-ZÁÉÍÓÚÀÃÕÂÊÎÔÛÇa-záéíóúàãõâêîôû\s\.\-]+:\s*$')
    hidden_re = re.compile(r'^exibir jogos \(\d+\)$')
    score_re  = re.compile(r'^\d+$')
    date_re   = re.compile(r'^\d{2}/\d{2}')

    linhas = [l.strip() for l in texto.split('\n')]
    jogos = []
    liga_atual = ''
    i = 0
    while i < len(linhas):
        linha = linhas[i]
        if (not linha or linha in SKIP or hidden_re.match(linha) or score_re.match(linha) or date_re.match(linha)):
            i += 1
            continue
        if time_re.match(linha):
            times = []
            j = i + 1
            while j < len(linhas) and len(times) < 2:
                cand = linhas[j].strip()
                if (not cand or cand in SKIP or cand == 'SRF' or score_re.match(cand) or hidden_re.match(cand)):
                    j += 1
                    continue
                if time_re.match(cand) or pais_re.match(cand): break
                times.append(cand)
                j += 1
            if len(times) == 2 and liga_atual:
                jogos.append({'h': linha, 'e': esporte, 'l': liga_atual, 'm': times[0], 'v': times[1]})
            i += 1
            continue
        if pais_re.match(linha):
            i += 1
            continue
        for j in range(i + 1, min(i + 4, len(linhas))):
            prox = linhas[j].strip()
            if not prox: continue
            if pais_re.match(prox) or prox in {'Classificação', 'Tabela', 'Classificação ao vivo'}:
                liga_atual = linha
            break
        i += 1
    return jogos

def extrair_todos_jogos_com_ia(dados_por_esporte):
    todos_jogos = []
    for esporte, texto in dados_por_esporte.items():
        todos_jogos.extend(extrair_jogos_esporte_direto(esporte, texto))
    return todos_jogos

# =============================================
# 3. CÁLCULOS
# =============================================
def calcular_heatmap(jogos):
    janelas = [(datetime.strptime("00:00", "%H:%M") + timedelta(minutes=30 * i)).strftime("%H:%M") for i in range(48)]
    heatmap = []
    for janela in janelas:
        h_janela = datetime.strptime(janela, "%H:%M")
        c = {"janela": janela, "Futebol": 0, "Basquete": 0, "Tenis": 0, "Hoquei": 0, "Futebol Americano": 0, "Beisebol": 0, "total": 0}
        for jogo in jogos:
            try:
                h_jogo = datetime.strptime(jogo['h'], "%H:%M")
                e = jogo['e']
                h_fim = h_jogo + timedelta(minutes=DURACAO.get(e, 105))
                if h_jogo <= h_janela < h_fim:
                    if e in c: c[e] += 1
                    c["total"] += 1
            except: continue
        heatmap.append(c)
    return heatmap

def calcular_power_hours(heatmap): 
    return sorted(heatmap, key=lambda x: x['total'], reverse=True)[:3]

def calcular_alertas(jogos, power_hours):
    alertas = []
    janelas_power = [p['janela'] for p in power_hours]
    for jogo in jogos:
        liga = jogo.get('l', '').lower()
        if not any(lp in liga for lp in LIGAS_PRIORITARIAS): continue
        try:
            h_jogo = datetime.strptime(jogo['h'], "%H:%M")
            h_fim = h_jogo + timedelta(minutes=DURACAO.get(jogo['e'], 105))
            for janela in janelas_power:
                h_j = datetime.strptime(janela, "%H:%M")
                if h_jogo <= h_j < h_fim:
                    alertas.append({"janela": janela, "jogo": f"{jogo['m']} x {jogo['v']}", "liga": jogo['l'], "esporte": jogo['e']})
        except: continue
    return alertas

# =============================================
# 4. EXCEL MELHORADO (MAPA DE CALOR SEMÁFORO FORTE)
# =============================================
def formatar_aba_resumo(ws):
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
    ws.row_dimensions[1].height = 20

def formatar_aba_heatmap(ws):
    ws.column_dimensions['A'].width = 12
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']: 
        ws.column_dimensions[col].width = 15

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    regra_cor = ColorScaleRule(
        start_type='min', start_color='00B050',       # Verde Forte e Vibrante
        mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Amarelinho claro 
        end_type='max', end_color='FF0000'            # Vermelho Puro 
    )
    ws.conditional_formatting.add(f"B2:H{ws.max_row}", regra_cor)

def formatar_aba_power(ws):
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="FFFF6600", end_color="FFFF6600", fill_type="solid")

def gerar_excel(jogos, heatmap, power_hours, alertas, data: datetime = None):
    if data is None: data = datetime.now()
    filename = f"Analise_Surebet_{data.strftime('%d_%m_%Y')}.xlsx"
    
    resumo = {
        "Info": [
            "📅 Data Analisada", "🎮 Total de Jogos", 
            "⚽ Futebol", "🏀 Basquete", "🎾 Tênis", "🏒 Hóquei", "🏈 Fut. Americano", "⚾ Beisebol", 
            "🔥 Pico 1", "🔥 Pico 2", "🔥 Pico 3"
        ],
        "Valor": [
            data.strftime('%d/%m/%Y'), len(jogos), 
            sum(1 for j in jogos if j['e'] == 'Futebol'), 
            sum(1 for j in jogos if j['e'] == 'Basquete'), 
            sum(1 for j in jogos if j['e'] == 'Tenis'), 
            sum(1 for j in jogos if j['e'] == 'Hoquei'), 
            sum(1 for j in jogos if j['e'] == 'Futebol Americano'), 
            sum(1 for j in jogos if j['e'] == 'Beisebol'), 
            f"{power_hours[0]['janela']} ({power_hours[0]['total']} jogos)" if len(power_hours) > 0 else "-", 
            f"{power_hours[1]['janela']} ({power_hours[1]['total']} jogos)" if len(power_hours) > 1 else "-", 
            f"{power_hours[2]['janela']} ({power_hours[2]['total']} jogos)" if len(power_hours) > 2 else "-"
        ]
    }
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        pd.DataFrame(resumo).to_excel(writer, sheet_name='Resumo', index=False)
        df_heat = pd.DataFrame(heatmap)
        df_heat.columns = ["Janela", "Futebol", "Basquete", "Tênis", "Hóquei", "Fut. Americano", "Beisebol", "Total"]
        df_heat.to_excel(writer, sheet_name='Heatmap', index=False)
        
        if alertas:
            df_al = pd.DataFrame(alertas)
            df_al.columns = ["Janela", "Jogo", "Liga", "Esporte"]
            df_al.to_excel(writer, sheet_name='Jogos Power Hour', index=False)
        else:
            pd.DataFrame({"Aviso": ["Nenhum jogo prioritário"]}).to_excel(writer, sheet_name='Jogos Power Hour', index=False)
    
    wb = load_workbook(filename)
    formatar_aba_resumo(wb['Resumo'])
    formatar_aba_heatmap(wb['Heatmap'])
    formatar_aba_power(wb['Jogos Power Hour'])
    wb.save(filename)
    return filename

# =============================================
# 5. TELEGRAM INTERNALS
# =============================================
def enviar_mensagem(texto):
    requests.post(f"https://api.telegram.org/bot{TOKEN_TELEGRAM}/sendMessage", data={"chat_id": CHAT_ID, "text": texto, "parse_mode": "Markdown"})

def enviar_arquivo(arquivo):
    with open(arquivo, "rb") as f: requests.post(f"https://api.telegram.org/bot{TOKEN_TELEGRAM}/sendDocument", data={"chat_id": CHAT_ID}, files={"document": f})

async def rodar_analise(data: datetime):
    enviar_mensagem(f"⏳ *Kairós Engine:* Iniciando varredura estratégica para {data.strftime('%d/%m/%Y')}...")
    try:
        dados = await extrair_jogos_flashscore(data)
        if not dados: return
        jogos = extrair_todos_jogos_com_ia(dados)
        if not jogos: return
        heatmap = calcular_heatmap(jogos)
        power_hours = calcular_power_hours(heatmap)
        arquivo = gerar_excel(jogos, heatmap, power_hours, calcular_alertas(jogos, power_hours), data)
        
        # MENSAGEM DO TELEGRAM ATUALIZADA COM TODOS OS ESPORTES
        total_jogos = len(jogos)
        futebol = sum(1 for j in jogos if j['e'] == 'Futebol')
        basquete = sum(1 for j in jogos if j['e'] == 'Basquete')
        tenis = sum(1 for j in jogos if j['e'] == 'Tenis')
        hoquei = sum(1 for j in jogos if j['e'] == 'Hoquei')
        fut_americano = sum(1 for j in jogos if j['e'] == 'Futebol Americano')
        beisebol = sum(1 for j in jogos if j['e'] == 'Beisebol')
        
        msg_telegram = (
            f"✅ *Análise Concluída!*\n"
            f"📅 Data: {data.strftime('%d/%m/%Y')}\n"
            f"🎮 Total: {total_jogos} jogos\n\n"
            f"⚽ Futebol: {futebol}\n"
            f"🏀 Basquete: {basquete}\n"
            f"🎾 Tênis: {tenis}\n"
            f"🏒 Hóquei: {hoquei}\n"
            f"🏈 Fut. Americano: {fut_americano}\n"
            f"⚾ Beisebol: {beisebol}\n\n"
            f"🔥 *Power Hours:*\n"
        )
        for p in power_hours:
            msg_telegram += f"🔥 {p['janela']} — {p['total']} jogos ativos\n"
            
        msg_telegram += f"\n📊 Planilha enviada abaixo!"
        
        enviar_mensagem(msg_telegram)
        enviar_arquivo(arquivo)
        
    except Exception as e: enviar_mensagem(f"❌ Erro: {e}")

# =============================================
# 6. NOVO MOTOR: SERVIDOR DE COMANDOS LOCAL (LOCALHOST E NUVEM)
# =============================================
async def processar_comando_painel(data_cmd):
    data_cmd = data_cmd.lower().strip()
    print(f"🔌 Sinal recebido do Painel Kairós: Varre data -> {data_cmd}")
    if data_cmd == "hoje": dt = datetime.now()
    elif data_cmd in ["amanha", "amanhã"]: dt = datetime.now() + timedelta(days=1)
    else:
        try: dt = datetime.strptime(data_cmd, "%d/%m/%Y")
        except: return
    await rodar_analise(dt)

class LocalCommandHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args): return
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Access-Control-Allow-Private-Network', 'true')
        self.end_headers()
    def do_GET(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Private-Network', 'true')
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        parsed_url = urlparse(self.path)
        if parsed_url.path == "/analisar":
            query = parse_qs(parsed_url.query)
            data_param = query.get('data', ['hoje'])[0]
            asyncio.run_coroutine_threadsafe(processar_comando_painel(data_param), o_loop)
            self.wfile.write(b'{"status": "iniciado"}')

def iniciar_servidor_local():
    # PUXA A PORTA DINÂMICA DA NUVEM (Ex: Render) OU USA 5000 SE FOR LOCAL
    porta = int(os.environ.get("PORT", 5000))
    # 0.0.0.0 PERMITE ACESSO EXTERNO NA NUVEM
    server = HTTPServer(('0.0.0.0', porta), LocalCommandHandler)
    print(f"🔌 Servidor de Integração Kairós rodando na porta {porta}!")
    server.serve_forever()

# =============================================
# LOOP TELEGRAM (MANTIDO COMO BACKUP)
# =============================================
async def escutar_telegram():
    global o_loop
    o_loop = asyncio.get_running_loop()
    threading.Thread(target=iniciar_servidor_local, daemon=True).start()
    print("🤖 Bot Kairós Online e integrado ao Painel! Aguardando cliques...")
    ultimo_update = 0
    while True:
        try:
            url = f"https://api.telegram.org/bot{TOKEN_TELEGRAM}/getUpdates"
            resp = await o_loop.run_in_executor(None, lambda: requests.get(url, params={"offset": ultimo_update + 1, "timeout": 30}, timeout=35).json())
            for upd in resp.get("result", []):
                ultimo_update = upd["update_id"]
                texto = upd.get("message", {}).get("text", "").lower().strip()
                if "/analisar" in texto:
                    partes = texto.split(" ")
                    if len(partes) > 1:
                        cmd = partes[1].strip()
                        if cmd == "hoje": await rodar_analise(datetime.now())
                        elif cmd in ["amanha", "amanhã"]: await rodar_analise(datetime.now() + timedelta(days=1))
        except: pass
        await asyncio.sleep(1)

if __name__ == "__main__":
    asyncio.run(escutar_telegram())